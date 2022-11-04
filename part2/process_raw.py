from codecs import ignore_errors
from ipaddress import collapse_addresses
from lib2to3.pgen2.parse import ParseError
import pandas as pd
from openpyxl import load_workbook
import datetime
import numbers
import datetime

# clean up the excel and return as a list of dfs

# load sheet to a df
def sheet_to_df(filepath, sheet):
    wb = load_workbook(filepath)
    ws = wb[sheet]

    data = []

    for row in ws.values:
        data.append([item for item in row])

    df = pd.DataFrame(data[1:], columns=data[0])
    
    sheet = df

    return sheet

# drop nones and trim strings
def trim_df(df=None):
    if df is None:
        print("Dataframe must be passed")
    else:
        df.dropna(inplace=True,how='all', axis='columns')
        df.dropna(inplace=True,how='all')
        df.applymap(lambda x: x.strip() if isinstance(x,str) else x)

# handle cases where day is longer than month by decreasing day 
def handle_dates(df, col_name):
    for row in df[col_name]:
        if isinstance(row, numbers.Number): 
            df[col_name].mask(df[col_name] == row,datetime.date(1899, 12,30) + datetime.timedelta(days=row) , inplace=True)
        else:
            try:
                row.strftime('%Y')
            except Exception:
                df[col_name].mask(df[col_name] == row, str(row)[:8] + str(int(str(row)[8:10]) - 1) + str(row)[10:], inplace=True)

# check if a column is a ssno column
def check_is_ssNo(name):
    if 'ssno' not in name.lower() and 'social security' not in name.lower() and 'worker' not in name.lower() and 'patient' not in name.lower():
        return False
    else:
        return True

#check if a column is a date column
def check_is_date_col(df, col_name):
    years = [*range(1800, 2030, 1)]
    years = [str(x) for x in years]
    r = '|'.join(years)
    
    if 'date' in col_name.lower() or (df[col_name].astype(str).str.contains(r).all() and check_is_ssNo(col_name) is False):
        return True
    else:
        return False

# pass in as bool and handle exceptions
def detect_type(df=None):
    if df is None:
        print("Dataframe must be passed")
    else:
        for col in df.columns.values:
            if df[col].isin([1,0]).all():
                df[col] = df[col].astype(bool)
            else:
                if check_is_date_col(df, col) is True:
                    handle_dates(df, col)
                else:
                    try:
                        df[col] = df[col].apply(pd.to_numeric)
                    except TypeError:
                        df[col] = df[col].astype('string')
                    except ValueError:
                        df[col] = df[col].astype('string')
            
# reformat social security numbers
def process_ssNo(df):
    for col in df.columns.values:
        if check_is_ssNo(col) is True:
            df[['first','second']] = df[col].str.split('-', expand=True)
            
            if len(df[col][0]) == 13:
                    df['year'] = df['first'].str[2:4]
                    df['month'] = df['first'].str[4:6]
                    df['day'] = df['first'].str[6:]
            else:
                df['year'] = df['first'].str[:2]
                df['month'] = df['first'].str[2:4]
                df['day'] = df['first'].str[4:]
            
            df[col] = df['day']+ df['month'] + df['year'] + '-' + df['second']
    
            df= df.drop(['first','second', 'day', 'month', 'year'],axis=1,inplace=True)

    #print(df)        

# read whole excel into a dict with df
def read_whole_excel(filepath):
    wb = load_workbook(filepath)
    df_list ={}
    
    for sheet in wb.sheetnames:
        df_list[sheet] = (sheet_to_df(filepath, sheet))
    
    return df_list

# clean every dataframe we get from an excel
def clean_whole_excel(filepath):
    df_list = read_whole_excel(filepath)

    for df in df_list.values():
        trim_df(df)
        process_ssNo(df)
        detect_type(df)

    print("Data cleaned.")
    
    return df_list

#